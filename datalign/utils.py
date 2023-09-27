import os
import pandas as pd
import re


def custom_read_df(path, sep=",", skiprows=0, lowerall=False):
    ext = path.split(".")[-1]
    encodings = ["utf-8", "ISO-8859-1", "cp1252", "latin1"]
    for encoding in encodings:
        try:
            if ext in ["xlsx", "xls"]:
                tmp = pd.read_excel(
                    path,
                    nrows=2,
                    skiprows=skiprows,
                    keep_default_na=False,
                )
                df = pd.read_excel(
                    path,
                    keep_default_na=False,
                    skiprows=skiprows,
                    dtype={
                        k: str
                        for k in tmp.columns
                        if not pd.api.types.is_datetime64_any_dtype(tmp[k])
                    },
                )

            if ext in ["csv"]:
                tmp = pd.read_csv(
                    path,
                    sep=sep,
                    skiprows=skiprows,
                    low_memory=True,
                    nrows=10,
                    keep_default_na=False,
                    on_bad_lines="warn",
                    engine="python",
                )

                df = pd.read_csv(
                    path,
                    sep=sep,
                    skiprows=skiprows,
                    low_memory=True,
                    dtype={
                        k: str
                        for k in tmp.columns
                        if not pd.api.types.is_datetime64_any_dtype(tmp[k])
                    },
                    keep_default_na=False,
                    on_bad_lines="warn",
                    engine="python",
                )
            break
        except UnicodeDecodeError:
            print(f"{encoding} failed")
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y/%m/%d")

    if lowerall:
        df.columns = df.columns.str.lower()
        df = df.apply(lambda x: x.astype(str).str.lower())

    return df


def get_transformations(sourcedf, workdir, prefix="", new_cols={}, lowerall=False):
    """
    get all transformations defined for a dataframe
    they should be located in transformations folder as csv files
    all csv files in the transformations should :
        follows this naming convention {prefix}{fieldname_in_df}.csv (prefix is optional)
        have ; as separator
        have to columns src and dest, dest is the value src has to be transformed into
        if a value is found but not present in src it is ignored and kept this way
    new_col is a dict of columns names that has to be created when applying transformation
    if column name is not found in new_col values will be replaced in the current column
    returns : the transformed dataframe
    """
    source = sourcedf.copy()
    print(f"tranformation working dir : {workdir}")
    print(source.columns.tolist())
    if os.path.isdir(os.path.join(workdir, "transformations")):
        for filename in os.listdir(os.path.join(workdir, "transformations")):
            fieldnames = filename.replace(prefix, "").replace(".csv", "")

            if lowerall:
                fieldnames = fieldnames.lower()
            print(f"Field to transform: {fieldnames}")
            if fieldnames in source.columns.tolist():
                trans_dict = (
                    custom_read_df(
                        os.path.join(workdir, "transformations", filename),
                        sep=";",
                        lowerall=lowerall,
                    )
                    .set_index("src")
                    .to_dict()["dest"]
                )
                print(f"\tTransforming {fieldnames}")
                if fieldnames not in new_cols.keys():
                    source[fieldnames] = source[fieldnames].apply(
                        lambda x: trans_dict[x] if x in trans_dict.keys() else x
                    )
                else:
                    source[new_cols[fieldnames]] = source[fieldnames].apply(
                        lambda x: trans_dict[x] if x in trans_dict.keys() else x
                    )
    return source


def recursive_lower(item):
    if item is None:
        return None
    elif isinstance(item, str):
        return item.lower()
    elif isinstance(item, list):
        return [recursive_lower(sub_item) for sub_item in item]
    else:
        raise ValueError("Unsupported data type")


def is_list_of_lists(variable):
    return (
        variable is not None
        and isinstance(variable, list)
        and all(isinstance(item, list) for item in variable)
    )


def excel_autoadjust_col(path, padding=5):
    import os
    import openpyxl
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path)
    sheets = [sheet for sheet in wb.get_sheet_names()]

    for sheet in sheets:
        ws = wb[sheet]
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            width = 0
            for row in range(ws.min_row, ws.max_row + 1):
                cell_value = ws.cell(column=col, row=row).value
                if cell_value:
                    cell_len = len(str(cell_value))
                    if cell_len > width:
                        width = cell_len + padding

            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=width
            )

        ws.column_dimensions = dim_holder

    wb.save(path)
    print("Completed adjustments for {}".format(os.path.basename(path)))


def remove_invalid_characters(input_string):
    return re.sub(r"[@\[\]/\\*\'\?[\]:]+", "", input_string)


def apply_comparison_formatting(df, writer, sheet_name):
    import xlsxwriter

    worksheet = writer.sheets[sheet_name]
    workbook = writer.book
    (max_row, max_col) = df.shape
    if max_row > 0:
        for i in range(max_col):
            col_name = xlsxwriter.utility.xl_col_to_name(i)
            if i % 2 == 0:
                col_near = xlsxwriter.utility.xl_col_to_name(i + 1)
            else:
                col_near = xlsxwriter.utility.xl_col_to_name(i - 1)

            crit_eq = (
                f"={col_name}2:{col_name}{max_row}={col_near}2:{col_near}{max_row}"
            )
            crit_dif = (
                f"={col_name}2:{col_name}{max_row}<>{col_near}2:{col_near}{max_row}"
            )

            worksheet.conditional_format(
                1,
                i,
                max_row,
                i,
                {
                    "type": "formula",
                    "criteria": crit_eq,
                    "format": get_workbookformat(workbook, kind="good"),
                },
            )
            worksheet.conditional_format(
                1,
                i,
                max_row,
                i,
                {
                    "type": "formula",
                    "criteria": crit_dif,
                    "format": get_workbookformat(workbook, kind="miss"),
                },
            )


def get_workbookformat(workbook, kind="good"):
    if kind == "perfect":
        return workbook.add_format({"bg_color": "#92D050"})
    if kind == "good":
        return workbook.add_format({"bg_color": "#56A4AC"})
    if kind == "wrong":
        return workbook.add_format({"bg_color": "#FB8F18"})
    if kind == "miss":
        return workbook.add_format({"bg_color": "#B33A3A"})
    if kind == "tofill":
        return workbook.add_format({"bg_color": "#FFFF00"})
    # FFFF00
